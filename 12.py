import tabula
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
import PyPDF2
import os
from datetime import datetime, timedelta
import time
import asyncio
import matplotlib.pyplot as plt
import sys
from PyInstaller import __main__ as pyi_main
from PIL import Image, ImageTk

# Changez le répertoire de travail au répertoire du script
os.chdir(os.path.dirname(os.path.abspath(__file__)))

if 'frozen' in getattr(sys, 'frozen', ''):
    # Supprimer les fichiers binaires en double lors de l'utilisation de PyInstaller
    a = pyi_main.run.pyi_makespec(pyi_main.run.PyiOptions())
    for b in a.binaries.copy():
        for d in a.datas:
            if b[1].endswith(d[0]):
                a.binaries.remove(b)
                break
    pyi_main.run.pymakespec.main(args)


# Définir des variables globales pour le dossier des PDFs, le fichier correct_answers et le dossier des excels
feuilles_corrigees = 0
folder_path = ""
correct_answers_file = ""
excel_folder = "QCM_excels"

# Fonction pour compter le nombre de fichiers PDF dans un dossier
def count_pdfs_in_folder(folder):
    pdf_count = sum(1 for file in os.listdir(folder) if file.lower().endswith('.pdf'))
    return pdf_count

# Fonction pour afficher le nombre total de fichiers PDF dans la console
def display_pdf_count():
    if folder_path:
        pdf_count = count_pdfs_in_folder(folder_path)
        print(f"Nombre total de fichiers PDF : {pdf_count}")
        # Ajouter une ligne pour afficher le nombre dans l'interface graphique
        pdf_count_label.config(text=f"Nombre total de fichiers PDF : {pdf_count}")
    else:
        print("Sélectionnez d'abord le dossier des PDFs.")
        pdf_count_label.config(text="Sélectionnez d'abord le dossier des PDFs.")

# Fonction pour extraire les tables des PDFs
async def extract_tables_from_pdfs():
    tables_list = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            file_path = os.path.join(folder_path, filename)
            tables = tabula.read_pdf(file_path, pages="all")
            for table in tables:
                tables_list.append(table)
    return tables_list

# Fonction pour enregistrer les tables dans des fichiers Excel
def save_tables_as_excel(tables_list):
    for i, table in enumerate(tables_list):
        file_path = os.path.join(excel_folder, 'QCM_table{}.xlsx'.format(i))
        table.to_excel(file_path, index=False)

# Fonction pour comparer les réponses entre deux fichiers Excel et compter le score final
def correct_excel_quiz(student_file, correct_answers_file):
    student_answers = pd.read_excel(student_file)
    answers = pd.read_excel(correct_answers_file)
    correct_count = 0
    for i, row in student_answers.iterrows():
        answer = row['Answer']
        if answer == answers.iloc[i]['Correct Answer']:
            correct_count += 1
    final_score = correct_count
    return final_score

def correct_quiz():
    if not folder_path or not correct_answers_file:
        print("Sélectionnez d'abord le dossier des PDFs et le fichier correct_answers.")
        return
    tables_list = asyncio.run(extract_tables_from_pdfs())
    save_tables_as_excel(tables_list)
    print("✅ Tables extraites et sauvegardées avec succès.")
    print("✅ Quiz corrigé avec succès.")

# Fonction pour extraire les noms et les notes et les stocker dans un fichier Excel
def extract_data_and_store_results():
    notes = []
    names = []
    for filename in os.listdir(excel_folder):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(excel_folder, filename)
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            cell = worksheet["C2"]
            cell.value
            final_score = correct_excel_quiz(file_path, correct_answers_file)
            new_name = cell.value
            names.append(new_name)
            new_note = final_score
            final_score_formatted = "{0:.2f}".format(final_score)
            notes.append(final_score_formatted)

    store_results_in_excel(names, notes)
    print("✅ Stockage de résultats avec succès")  # Print success message

# Fonction pour stocker les résultats dans un fichier Excel
def store_results_in_excel(names, notes):
    if not os.path.exists("resultats.xlsx"):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "Noms"
        worksheet["B1"] = "Notes"
        workbook.save("resultats.xlsx")

    df = pd.read_excel("resultats.xlsx")
    new_data = {"Noms": names, "Notes": notes}
    new_df = pd.DataFrame(data=new_data)
    df = pd.concat([df, new_df], ignore_index=True)
    df.to_excel("resultats.xlsx", index=False)

# Fonction pour afficher les résultats
def show_results():
    os.system("xdg-open resultats.xlsx")

def show_statistics():
    statistics_window = tk.Toplevel(root)
    statistics_window.title("Statistics")

    # Bouton pour afficher les statistiques
    statistics_button = tk.Button(statistics_window, text="Generate Statistics", command=generate_statistics)
    statistics_button.pack()

def generate_statistics():
    # Charger les données depuis le fichier Excel
    df = pd.read_excel("resultats.xlsx")

    # Diviser les étudiants en cinq groupes en fonction des intervalles de notes
    bins = [0, 5, 10, 15, 18, 20]
    labels = ['0-5', '6-10', '11-15', '16-18', '19-20']
    df['NoteGroup'] = pd.cut(df['Notes'], bins=bins, labels=labels)

    # Compter le nombre d'étudiants dans chaque groupe
    group_counts = df['NoteGroup'].value_counts()

    # Définir une palette de couleurs personnalisée
    colors = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99', '#c2c2f0']

    # Trouver la meilleure et la moins bonne note avec les noms correspondants
    best_student_info = df.loc[df['Notes'].idxmax(), ['Noms', 'Notes']]
    worst_student_info = df.loc[df['Notes'].idxmin(), ['Noms', 'Notes']]

    # Créer un graphique en secteurs (pie chart) avec des couleurs personnalisées
    plt.figure(figsize=(8, 8))
    wedges, texts, autotexts = plt.pie(group_counts, labels=group_counts.index, autopct='%1.1f%%', startangle=140, colors=colors, textprops=dict(color="w"))

    # Ajouter une légende (Clé illustrée)
    plt.legend(wedges, labels, title="Intervalles de notes", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Ajouter la meilleure et la moins bonne note avec les noms correspondants à la titre
    plt.title(f"Autocorrect_QCM result_graph\nMeilleure note: {best_student_info['Notes']} ({best_student_info['Noms']})\nMoins bonne note: {worst_student_info['Notes']} ({worst_student_info['Noms']})")

    # Afficher le graphique
    plt.show()
    create_pie_chart_with_legend_and_extremes("resultats.xlsx")


# Fonction pour créer l'interface graphique
# Fonction pour créer l'interface graphique
def create_gui():
    global root
    root = tk.Tk()
    root.title("QCM Application")

    # Ajouter un widget Text pour afficher la console au centre
    global console_text
    console_text = tk.Text(root, height=10, width=50, wrap=tk.WORD)
    console_text.grid(row=0, column=0, columnspan=4, pady=10)

    # Rediriger la sortie de la console vers le widget Text
    sys.stdout = ConsoleRedirector(console_text)

    # Chemin vers l'image (remplacez 'chemin/vers/votre/image.png' par le chemin de votre image)
    image_path = '4.png'

    # Chargez l'image avec Pillow
    image = Image.open(image_path)
    tk_image = ImageTk.PhotoImage(image)
    
    # Bouton 1 : Start
    start_button = tk.Button(root, text="Start", command=start_application)
    start_button.grid(row=1, column=0, pady=0, sticky="ew")

    # Bouton 2 : Sélectionner QCM_pdfs
    select_pdfs_button = tk.Button(root, text="Sélectionner QCM_pdfs", command=select_pdfs_folder)
    select_pdfs_button.grid(row=2, column=0, pady=0, sticky="ew")

    # Bouton 3 : Sélectionner QCM_correct.xlsx
    select_correct_file_button = tk.Button(root, text="Sélectionner QCM_correct.xlsx ", command=select_correct_file)
    select_correct_file_button.grid(row=3, column=0, pady=0, sticky="ew")

    # Bouton 4 : Corriger
    correct_button = tk.Button(root, text="Corriger", command=correct_quiz)
    correct_button.grid(row=4, column=0, pady=0, sticky="ew")

    # Bouton 5 : Stockage de résultats
    store_results_button = tk.Button(root, text="Stockage de résultats", command=extract_data_and_store_results)
    store_results_button.grid(row=5, column=0, pady=0, sticky="ew")

    # Bouton 6 : Afficher résultats
    show_results_button = tk.Button(root, text="Afficher résultats", command=show_results)
    show_results_button.grid(row=6, column=0, pady=0, sticky="ew")

    # Bouton 8 : Statistics
    statistics_button = tk.Button(root, text="Statistics", command=generate_statistics)
    statistics_button.grid(row=7, column=0, pady=0, sticky="ew")

    # Bouton 7 : Quitter
    quit_button = tk.Button(root, text="Quitter", command=quit_application)
    quit_button.grid(row=8, column=0, pady=0, sticky="ew")

    image_label = tk.Label(root, image=tk_image)
    image_label.grid(row=1, column=1, rowspan=8, padx=10, pady=0, sticky="w")

    root.mainloop()

# Classe pour rediriger la sortie de la console vers un widget Text
class ConsoleRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, text):
        self.text_widget.insert(tk.END, text)
        self.text_widget.see(tk.END)

# Fonction pour quitter l'application
def quit_application():
    print("✅ Good bye!")  # Print farewell message
    if root:
        root.after(1000, root.destroy)  # Destroy the Tkinter root window if it exists

# Fonction pour démarrer l'application (à personnaliser selon vos besoins)
def start_application():
    print("✅ L'application a démarré.")

# Fonction pour sélectionner le dossier des PDFs
def select_pdfs_folder():
    global folder_path
    folder_path = filedialog.askdirectory(title="Sélectionnez le dossier des PDFs")
    folder_path = os.path.join(folder_path, "QCM_pdfs")
    print("✅ Dossier des PDFs sélectionné : ", folder_path)
    display_pdf_count()

# Fonction pour sélectionner le fichier correct_answers
def select_correct_file():
    global correct_answers_file
    correct_answers_file = filedialog.askopenfilename(title="Sélectionnez QCM_correct.xlsx")
    print("✅ Fichier QCM_correct.xlsx sélectionné : ", correct_answers_file)

# Fonction pour extraire les noms et les notes et les stocker dans un fichier Excel
def extract_data_and_store_results():
    notes = []
    names = []
    for filename in os.listdir(excel_folder):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(excel_folder, filename)
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            cell = worksheet["C2"]
            cell.value
            final_score = correct_excel_quiz(file_path, correct_answers_file)
            new_name = cell.value
            names.append(new_name)
            new_note = final_score
            final_score_formatted = "{0:.2f}".format(final_score)
            notes.append(final_score_formatted)

    store_results_in_excel(names, notes)
    print("✅ Stockage de résultats avec succès")  # Print success message

# Lancer l'interface graphique
create_gui()
